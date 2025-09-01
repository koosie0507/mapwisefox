from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from pandas import ExcelWriter

from mapwisefox.web.model import Evidence


class PandasRepo:
    __MANDATORY_COLS = ["include", "exclude_reasons"]

    def __init__(
        self,
        excel_file: Path,
        sheet_name: Optional[str] = None,
        aliases: dict[str, str] = None,
    ):
        self._excel_file = excel_file
        self._sheet_name = sheet_name or self.__read_first_sheet_name()
        excel = pd.read_excel(excel_file, sheet_name=self._sheet_name, index_col=0, na_filter=False)
        assert isinstance(excel, pd.DataFrame)
        self._df: pd.DataFrame = excel.sort_index(axis=1, inplace=False)
        if aliases is not None:
            self._df.rename(columns=aliases, inplace=True)
        self._aliases = aliases
        self.__infer_cluster_id()
        self.__ensure_mandatory_cols()

    @property
    def dataframe(self) -> pd.DataFrame:
        return self._df

    def get(self, cluster_id: int) -> Evidence:
        params = self._df.loc[cluster_id].to_dict()
        params.update({
            "cluster_id": cluster_id,
        })
        return Evidence(**params)

    def update(self, evidence: Evidence) -> None:
        if evidence.cluster_id not in self._df.index:
            raise KeyError(evidence.cluster_id)
        self._df.loc[evidence.cluster_id] = evidence.model_dump()
        self.__write_xls()

    def __read_first_sheet_name(self) -> str:
        wb: Optional[Workbook] = None
        try:
            wb = load_workbook(filename=self._excel_file, read_only=True)
            return wb.sheetnames[0]
        finally:
            wb.close()

    def __infer_cluster_id(self) -> None:
        if "cluster_id" not in self._df.columns:
            self._df.index.set_names(["cluster_id"], inplace=True)

    def __ensure_mandatory_cols(self) -> None:
        for col in self.__MANDATORY_COLS:
            if col in self._df.columns:
                self._df[col] = self._df[col].astype(str)
            else:
                kw_args = {col: pd.Series([""]*len(self._df), dtype=str)}
                self._df = self._df.assign(**kw_args)

    def __write_xls(self):
        with ExcelWriter(self._excel_file, "openpyxl", mode="a", if_sheet_exists="replace") as writer:
            if isinstance(self._sheet_name, str):
                self._df.to_excel(writer, self._sheet_name, index=True, index_label="cluster_id", na_rep="")
            else:
                self._df.to_excel(writer, index=True, index_label="cluster_id", na_rep="")