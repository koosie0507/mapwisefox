import asyncio
from io import BytesIO

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook

from mapwisefox.web.config import AppSettings
from mapwisefox.web.api import auth_api_router, config_api_router, workbooks_api_router
from mapwisefox.web._deps import current_user, settings, user_upload_dir
from mapwisefox.web.model import UserInfo


HEADERS = [
    "title",
    "abstract",
    "doi",
    "authors",
    "keywords",
    "year",
    "source",
    "url",
    "has_pdf",
    "referencing_paper_ids",
]


@pytest.fixture
def workbook_file():
    stream = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Studies"
    worksheet.append(HEADERS)
    worksheet.append(
        [
            "Title",
            "Abstract",
            "10.1/test",
            "Ada",
            "test",
            2025,
            "Journal",
            "",
            False,
            "",
        ]
    )
    worksheet.append(
        [
            "Second",
            "Abstract",
            "10.1/second",
            "Grace",
            "test",
            2024,
            "Journal",
            "",
            False,
            "",
        ]
    )
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


@pytest.fixture
def client(tmp_path):
    app = FastAPI()
    app.include_router(auth_api_router)
    app.include_router(config_api_router)
    app.include_router(workbooks_api_router)
    config = AppSettings(uploads_dir=tmp_path)
    app.dependency_overrides[current_user] = lambda: None
    app.dependency_overrides[user_upload_dir] = lambda: tmp_path
    app.dependency_overrides[settings] = lambda: config
    return TestClient(app)


@pytest.fixture
def basic_upload(client, workbook_file):
    def _():
        return client.post(
            "/api/v1/workbooks",
            files={"file": ("studies.xlsx", workbook_file)},
            data={"worksheetName": "Studies", "expectedColumns": "title,abstract"},
        )

    return _


@pytest.fixture
def imported_client(client, workbook_file, basic_upload):
    response = basic_upload()
    assert response.status_code == 201
    return client


def test_auth_required_reports_server_configuration(client):
    response = client.get("/api/v1/auth/required")

    assert response.json() == {"required": False}


def test_auth_required_reports_enabled_configuration(client):
    config = client.app.dependency_overrides[settings]().model_copy(
        update={"auth_enabled": True}
    )
    client.app.dependency_overrides[settings] = lambda: config

    response = client.get("/api/v1/auth/required")

    assert response.json() == {"required": True}


def test_config_returns_frontend_context(client):
    response = client.get("/api/v1/config")

    assert response.json() == {
        "user": None,
        "worksheetName": "",
        "expectedColumns": "",
        "decisionColumn": "include",
        "exclusionReasonColumn": "exclude_reason",
    }


def test_config_returns_current_user(client):
    client.app.dependency_overrides[current_user] = lambda: UserInfo(
        dirname="ada-lovelace",
        issuer="https://identity.example.com",
        subject="ada",
        display_name="Ada Lovelace",
        email="ada@example.com",
    )

    response = client.get("/api/v1/config")

    assert response.json()["user"] == {
        "display_name": "Ada Lovelace",
        "email": "ada@example.com",
    }


def test_import_returns_resolved_resource(client, workbook_file):
    response = client.post(
        "/api/v1/workbooks",
        files={"file": ("studies.xlsx", workbook_file)},
        data={"worksheetName": "Studies", "expectedColumns": "title,abstract"},
    )

    assert response.status_code == 201
    assert response.json()["recordCount"] == 2
    assert response.json()["unfilledRecordCount"] == 2
    assert response.headers["location"] == "/api/v1/workbooks/studies.xlsx"


def test_import_rejects_unresolved_configuration(client, workbook_file):
    response = client.post(
        "/api/v1/workbooks",
        files={"file": ("studies.xlsx", workbook_file)},
    )

    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "missing_import_configuration"


def test_import_rejects_invalid_xlsx_content(client):
    response = client.post(
        "/api/v1/workbooks",
        files={"file": ("broken.xlsx", b"not an xlsx workbook")},
        data={"worksheetName": "Studies", "expectedColumns": "title"},
    )

    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "invalid_workbook"


def test_screening_get_uses_zero_based_index(imported_client):
    response = imported_client.get("/api/v1/workbooks/studies.xlsx/screening/0")

    assert response.status_code == 200
    assert response.json()["recordIndex"] == 0
    assert response.json()["decision"] == "undecided"


def test_screening_patch_persists_decision(imported_client, tmp_path):
    response = imported_client.patch(
        "/api/v1/workbooks/studies.xlsx/screening/0",
        json={"decision": "excluded", "exclusionReasons": ["Not Software"]},
    )

    workbook = load_workbook(tmp_path / "studies.xlsx")
    worksheet = workbook["Studies"]
    assert response.status_code == 200
    assert response.json()["exclusionReasons"] == ["not software"]
    assert worksheet["K2"].value == "exclude"
    assert worksheet["L2"].value == "not software"
    workbook.close()


def test_workbook_list_tracks_unfilled_records(imported_client):
    response = imported_client.patch(
        "/api/v1/workbooks/studies.xlsx/screening/0",
        json={"decision": "included", "exclusionReasons": []},
    )

    assert response.status_code == 200
    listed = imported_client.get("/api/v1/workbooks").json()

    assert listed[0]["recordCount"] == 2
    assert listed[0]["unfilledRecordCount"] == 1


def test_get_workbook_tracks_unfilled_records(imported_client):
    imported_client.patch(
        "/api/v1/workbooks/studies.xlsx/screening/0",
        json={"decision": "included", "exclusionReasons": []},
    )

    response = imported_client.get("/api/v1/workbooks/studies.xlsx")

    assert response.json()["unfilledRecordCount"] == 1


def test_workbook_list_reports_complete_workbook(imported_client):
    for index in range(2):
        response = imported_client.patch(
            f"/api/v1/workbooks/studies.xlsx/screening/{index}",
            json={"decision": "included", "exclusionReasons": []},
        )
        assert response.status_code == 200

    listed = imported_client.get("/api/v1/workbooks").json()

    assert listed[0]["unfilledRecordCount"] == 0


def test_reimport_without_prior_delete_refreshes_cached_unfilled_count(
    imported_client, basic_upload, workbook_file
):
    imported_client.patch(
        "/api/v1/workbooks/studies.xlsx/screening/0",
        json={"decision": "included", "exclusionReasons": []},
    )
    previous_unfilled = imported_client.get("/api/v1/workbooks").json()[0][
        "unfilledRecordCount"
    ]

    # no prior delete, just re-upload.
    current_unfilled = basic_upload().json()["unfilledRecordCount"]

    assert previous_unfilled == 1
    assert current_unfilled == 2


def test_reimport_with_prior_delete_refreshes_cached_unfilled_count(
    imported_client, basic_upload, workbook_file
):
    imported_client.patch(
        "/api/v1/workbooks/studies.xlsx/screening/0",
        json={"decision": "included", "exclusionReasons": []},
    )
    previous_unfilled = imported_client.get("/api/v1/workbooks").json()[0][
        "unfilledRecordCount"
    ]
    imported_client.delete("/api/v1/workbooks/studies.xlsx")

    current_unfilled = basic_upload().json()["unfilledRecordCount"]

    assert previous_unfilled == 1
    assert current_unfilled == 2


def test_screening_patch_rejects_excluded_without_reason(imported_client):
    response = imported_client.patch(
        "/api/v1/workbooks/studies.xlsx/screening/0",
        json={"decision": "excluded", "exclusionReasons": []},
    )

    assert response.status_code == 422


def test_delete_removes_resource(imported_client):
    response = imported_client.delete("/api/v1/workbooks/studies.xlsx")

    assert response.status_code == 204
    assert imported_client.get("/api/v1/workbooks/studies.xlsx").status_code == 404


@pytest.mark.anyio
async def test_concurrent_patches_preserve_both_updates(tmp_path, workbook_file):
    app = FastAPI()
    app.include_router(workbooks_api_router)
    app.dependency_overrides[user_upload_dir] = lambda: tmp_path
    app.dependency_overrides[settings] = lambda: AppSettings(uploads_dir=tmp_path)
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        imported = await client.post(
            "/api/v1/workbooks",
            files={"file": ("studies.xlsx", workbook_file)},
            data={"worksheetName": "Studies", "expectedColumns": "title"},
        )
        assert imported.status_code == 201
        responses = await asyncio.gather(
            client.patch(
                "/api/v1/workbooks/studies.xlsx/screening/0",
                json={"decision": "excluded", "exclusionReasons": ["first"]},
            ),
            client.patch(
                "/api/v1/workbooks/studies.xlsx/screening/1",
                json={"decision": "excluded", "exclusionReasons": ["second"]},
            ),
        )

    workbook = load_workbook(tmp_path / "studies.xlsx")
    assert [response.status_code for response in responses] == [200, 200]
    assert workbook["Studies"]["K2"].value == "exclude"
    assert workbook["Studies"]["K3"].value == "exclude"
    workbook.close()
