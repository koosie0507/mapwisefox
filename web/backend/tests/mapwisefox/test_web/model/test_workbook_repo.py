from copy import copy
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from mapwisefox.common.config import SelectionConfig
from mapwisefox.web.model import (
    WorkbookRepository,
    WorkbookValidationError,
    metadata_path,
)


HEADERS = [
    "title",
    "abstract",
    "doi",
    "authors",
    "keywords",
    "year",
    "source",
    "url",
    "has_pdf",
    "referencing_paper_ids",
    "notes",
]
ROW = [
    "A title",
    "An abstract",
    "10.1/example",
    "Ada;Grace",
    "testing;software",
    2024,
    "Journal",
    "https://example.test",
    False,
    "",
    "preserve me",
]


@pytest.fixture
def selection_criteria() -> SelectionConfig:
    return SelectionConfig(
        review_topic="entity resolution",
        inclusion_criteria=[{"label": "english", "description": "written in English"}],
        exclusion_criteria=[{"label": "not software", "description": "no software"}],
    )


@pytest.fixture
def selection_config_path(tmp_path: Path) -> Path:
    path = tmp_path / "study-selection-config.json"
    path.write_text(
        '{"review_topic":"entity resolution",'
        '"inclusion_criteria":[{"label":"english","description":"written in English"}],'
        '"exclusion_criteria":[{"label":"not software","description":"no software"}]}'
    )
    return path


@pytest.fixture
def source_workbook(tmp_path):
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Studies"
    worksheet.append(HEADERS)
    worksheet.append(ROW)
    worksheet.append(["Second", *ROW[1:]])
    worksheet["A2"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
    # add a second sheet that must not be touched
    calculations = workbook.create_sheet("Calculations")
    calculations["A1"] = "=1+1"
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def imported_workbook(tmp_path, source_workbook):
    destination = tmp_path / "screening.xlsx"
    metadata = WorkbookRepository.import_workbook(
        source_workbook,
        destination,
        "Studies",
        {},
        "decision",
        "reason",
    )
    return destination, metadata


def test_import_persists_resolved_metadata(imported_workbook):
    destination, metadata = imported_workbook

    persisted = WorkbookRepository.read_metadata(destination)

    assert persisted == metadata


def test_import_appends_missing_screening_columns(imported_workbook):
    destination, _ = imported_workbook

    workbook = load_workbook(destination)
    headers = [cell.value for cell in workbook["Studies"][1]]
    workbook.close()

    assert headers[-2:] == ["decision", "reason"]


def test_import_rejects_missing_mandatory_mapped_field(source_workbook, tmp_path):
    with pytest.raises(WorkbookValidationError, match="Missing") as error:
        WorkbookRepository.import_workbook(
            source_workbook,
            tmp_path / "output.xlsx",
            "Studies",
            {"title": "missing"},
            "decision",
            "reason",
        )

    assert error.value.code == "missing_mandatory_fields"


def test_import_rejects_screening_column_mapped_to_evidence(source_workbook, tmp_path):
    with pytest.raises(WorkbookValidationError) as error:
        WorkbookRepository.import_workbook(
            source_workbook,
            tmp_path / "output.xlsx",
            "Studies",
            {"title": "title"},
            "title",
            "reason",
        )

    assert error.value.code == "screening_evidence_column_collision"


def test_import_uses_aliased_field_mapping(source_workbook, tmp_path):
    workbook = load_workbook(source_workbook)
    worksheet = workbook["Studies"]
    worksheet["I1"] = "full_text"
    worksheet["I2"] = "yes"
    workbook.save(source_workbook)
    workbook.close()
    destination = tmp_path / "output.xlsx"
    metadata = WorkbookRepository.import_workbook(
        source_workbook,
        destination,
        "Studies",
        {"hasPdf": "full_text"},
        "decision",
        "reason",
    )

    assert WorkbookRepository(destination, metadata).get(0).evidence.has_pdf is True


def test_import_rejects_internal_blank_row(source_workbook, tmp_path):
    workbook = load_workbook(source_workbook)
    worksheet = workbook["Studies"]
    worksheet.insert_rows(3)
    workbook.save(source_workbook)
    workbook.close()

    with pytest.raises(WorkbookValidationError) as error:
        WorkbookRepository.import_workbook(
            source_workbook,
            tmp_path / "output.xlsx",
            "Studies",
            {},
            "decision",
            "reason",
        )

    assert error.value.code == "blank_record_row"


def test_update_preserves_unrelated_workbook_content(imported_workbook):
    destination, metadata = imported_workbook
    before = load_workbook(destination)
    original_fill = copy(before["Studies"]["A2"].fill)
    before.close()

    WorkbookRepository(destination, metadata).update(0, "excluded", ["not software"])

    after = load_workbook(destination, data_only=False)
    assert after["Studies"]["K2"].value == "preserve me"
    assert after["Studies"]["A2"].fill == original_fill
    assert after["Calculations"]["A1"].value == "=1+1"
    after.close()


def test_update_changes_only_screening_values(imported_workbook):
    destination, metadata = imported_workbook

    record = WorkbookRepository(destination, metadata).update(
        0, "excluded", ["not software", "not english"]
    )

    assert record.decision == "excluded"
    assert record.exclusion_reasons == ["not software", "not english"]


def test_delete_removes_workbook_metadata(imported_workbook):
    destination, metadata = imported_workbook

    WorkbookRepository(destination, metadata).delete()

    assert not destination.exists()
    assert not metadata_path(destination).exists()


def test_import_persists_selection_criteria(
    tmp_path, source_workbook, selection_criteria
):
    destination = tmp_path / "with-criteria.xlsx"

    metadata = WorkbookRepository.import_workbook(
        source_workbook,
        destination,
        "Studies",
        {},
        "decision",
        "reason",
        selection_criteria,
    )

    assert metadata.selection_criteria == selection_criteria
    persisted = WorkbookRepository.read_metadata(destination)
    assert persisted.selection_criteria == selection_criteria


def test_import_omits_selection_criteria_by_default(imported_workbook):
    destination, metadata = imported_workbook

    assert metadata.selection_criteria is None
    persisted = WorkbookRepository.read_metadata(destination)
    assert persisted.selection_criteria is None


def test_import_round_trips_selection_criteria_in_sidecar(
    tmp_path, source_workbook, selection_criteria
):
    destination = tmp_path / "round-trip.xlsx"

    WorkbookRepository.import_workbook(
        source_workbook,
        destination,
        "Studies",
        {},
        "decision",
        "reason",
        selection_criteria,
    )

    sidecar = metadata_path(destination).read_text(encoding="utf-8")

    assert "selectionCriteria" in sidecar
    assert "not software" in sidecar
