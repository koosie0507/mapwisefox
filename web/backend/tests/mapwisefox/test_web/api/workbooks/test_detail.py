import asyncio

import pytest
from openpyxl import load_workbook


def test_get_workbook_tracks_unfilled_records(imported_client):
    imported_client.patch(
        "/api/v1/workbooks/studies.xlsx/screening/0",
        json={"decision": "included", "exclusionReasons": []},
    )

    response = imported_client.get("/api/v1/workbooks/studies.xlsx")

    assert response.json()["unfilledRecordCount"] == 1


def test_screening_get_uses_zero_based_index(imported_client):
    response = imported_client.get("/api/v1/workbooks/studies.xlsx/screening/0")

    assert response.status_code == 200
    assert response.json()["recordIndex"] == 0
    assert response.json()["decision"] == "undecided"


def test_screening_patch_persists_decision(imported_client, tmp_path):
    response = imported_client.patch(
        "/api/v1/workbooks/studies.xlsx/screening/0",
        json={"decision": "excluded", "exclusionReasons": ["Not Software"]},
    )

    workbook = load_workbook(tmp_path / "studies.xlsx")
    worksheet = workbook["Studies"]
    assert response.status_code == 200
    assert response.json()["exclusionReasons"] == ["not software"]
    assert worksheet["K2"].value == "exclude"
    assert worksheet["L2"].value == "not software"
    workbook.close()


def test_screening_patch_rejects_excluded_without_reason(imported_client):
    response = imported_client.patch(
        "/api/v1/workbooks/studies.xlsx/screening/0",
        json={"decision": "excluded", "exclusionReasons": []},
    )

    assert response.status_code == 422


def test_delete_removes_resource(imported_client):
    response = imported_client.delete("/api/v1/workbooks/studies.xlsx")

    assert response.status_code == 204
    assert imported_client.get("/api/v1/workbooks/studies.xlsx").status_code == 404


@pytest.mark.anyio
async def test_concurrent_patches_preserve_both_updates(
    async_client, tmp_path, workbook_file
):
    imported = await async_client.post(
        "/api/v1/workbooks",
        files={"file": ("studies.xlsx", workbook_file)},
        data={"worksheetName": "Studies", "expectedColumns": "title"},
    )
    assert imported.status_code == 201
    responses = await asyncio.gather(
        async_client.patch(
            "/api/v1/workbooks/studies.xlsx/screening/0",
            json={"decision": "excluded", "exclusionReasons": ["first"]},
        ),
        async_client.patch(
            "/api/v1/workbooks/studies.xlsx/screening/1",
            json={"decision": "excluded", "exclusionReasons": ["second"]},
        ),
    )

    workbook = load_workbook(tmp_path / "studies.xlsx")
    assert [response.status_code for response in responses] == [200, 200]
    assert workbook["Studies"]["K2"].value == "exclude"
    assert workbook["Studies"]["K3"].value == "exclude"
    workbook.close()
