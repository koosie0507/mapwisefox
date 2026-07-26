import json
import os
import tempfile
from zipfile import BadZipFile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ConfigDict, Field

from mapwisefox.common.config import SelectionConfig
from mapwisefox.web.model._evidence import Evidence


type Decision = Literal["undecided", "included", "excluded"]

LIST_SEPARATOR = ";"
EXCLUSION_REASON_COLUMN_ALIASES = ("exclude_reason", "exclude_reasons")
DECISION_VALUES: dict[str, Decision] = {
    "": "undecided",
    "include": "included",
    "exclude": "excluded",
}
PERSISTED_DECISIONS: dict[Decision, str] = {
    "undecided": "",
    "included": "include",
    "excluded": "exclude",
}


class WorkbookMetadata(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    name: str
    worksheet_name: str = Field(alias="worksheetName")
    header_row: int = Field(alias="headerRow")
    field_mappings: dict[str, str] = Field(default_factory=dict, alias="fieldMappings")
    decision_column: str = Field(alias="decisionColumn")
    exclusion_reason_column: str = Field(alias="exclusionReasonColumn")
    record_count: int = Field(alias="recordCount")
    unfilled_record_count: int = Field(default=0, alias="unfilledRecordCount")
    selection_criteria: SelectionConfig | None = Field(
        default=None, alias="selectionCriteria"
    )


@dataclass(frozen=True)
class ScreeningRecord:
    evidence: Evidence
    decision: Decision
    exclusion_reasons: list[str]


class WorkbookValidationError(ValueError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        row: int | None = None,
        column: str | None = None,
    ):
        super().__init__(message)
        self.code = code
        self.row = row
        self.column = column

    def detail(self) -> dict[str, str | int]:
        result: dict[str, str | int] = {"code": self.code, "message": str(self)}
        if self.row is not None:
            result["row"] = self.row
        if self.column is not None:
            result["column"] = self.column
        return result


def workbook_path(upload_dir: Path, name: str) -> Path:
    if Path(name).name != name or Path(name).suffix.lower() != ".xlsx":
        raise WorkbookValidationError(
            "invalid_name", "Workbook name must be a safe .xlsx filename"
        )
    return upload_dir / name


def metadata_path(path: Path) -> Path:
    return path.with_suffix(".metadata.json")


def _is_blank(value: Any) -> bool:
    return value is None or isinstance(value, str) and not value.strip()


def _headers(worksheet) -> tuple[int, list[str]]:
    for row in worksheet.iter_rows():
        values = [cell.value for cell in row]
        if all(_is_blank(value) for value in values) or len(row) < 1:
            continue
        last = max(index for index, value in enumerate(values) if not _is_blank(value))
        return row[0].row, _validate_headers(values[: last + 1], row[0].row)
    raise WorkbookValidationError("missing_header", "Worksheet contains no header row")


def _validate_headers(values: list[Any], row: int) -> list[str]:
    if any(_is_blank(value) or not isinstance(value, str) for value in values):
        raise WorkbookValidationError(
            "invalid_header", "Headers must be non-empty strings", row=row
        )
    headers = [value.strip() for value in values]
    duplicates = sorted({value for value in headers if headers.count(value) > 1})
    if duplicates:
        raise WorkbookValidationError(
            "duplicate_header", f"Duplicate headers: {', '.join(duplicates)}", row=row
        )
    return headers


def _resolve_exclusion_reason_column(headers: list[str], column: str) -> str:
    if column not in EXCLUSION_REASON_COLUMN_ALIASES or column in headers:
        return column
    return next(
        (alias for alias in EXCLUSION_REASON_COLUMN_ALIASES if alias in headers), column
    )


def _exclusion_reason_value(values: dict[str, Any], column: str) -> Any:
    columns = (
        (column, *EXCLUSION_REASON_COLUMN_ALIASES)
        if column in EXCLUSION_REASON_COLUMN_ALIASES
        else (column,)
    )
    return next(
        (values.get(name) for name in columns if not _is_blank(values.get(name))), None
    )


def _record_count(worksheet, header_row: int, width: int) -> int:
    rows = range(header_row + 1, worksheet.max_row + 1)
    populated = [
        row
        for row in rows
        if any(
            not _is_blank(worksheet.cell(row, column).value)
            for column in range(1, width + 1)
        )
    ]
    if not populated:
        raise WorkbookValidationError(
            "missing_records", "Worksheet contains no records"
        )
    last_row = populated[-1]
    blank_rows = [
        row for row in range(header_row + 1, last_row + 1) if row not in populated
    ]
    if blank_rows:
        raise WorkbookValidationError(
            "blank_record_row",
            "Blank rows are not allowed between records",
            row=blank_rows[0],
        )
    return last_row - header_row


def _write_json_atomic(path: Path, value: dict[str, Any]) -> None:
    descriptor, temporary_name = tempfile.mkstemp(dir=path.parent, suffix=".json")
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as stream:
            json.dump(value, stream, indent=2)
        os.replace(temporary_name, path)
    finally:
        Path(temporary_name).unlink(missing_ok=True)


class WorkbookRepository:
    def __init__(self, path: Path, metadata: WorkbookMetadata | None = None):
        self.path = path
        self.metadata = metadata or self.read_metadata(path)

    def __hash__(self):
        return hash(self.path)

    def __repr__(self):
        return self.path

    def __eq__(self, other):
        if not isinstance(other, self.__class__):
            return False
        return self.path == other.path

    @staticmethod
    def read_metadata(path: Path) -> WorkbookMetadata:
        sidecar = metadata_path(path)
        if not path.is_file() or not sidecar.is_file():
            raise FileNotFoundError(path.name)
        return WorkbookMetadata.model_validate_json(sidecar.read_text(encoding="utf-8"))

    @classmethod
    def import_workbook(
        cls,
        source: Path,
        destination: Path,
        worksheet_name: str | None,
        field_mappings: dict[str, str],
        decision_column: str,
        exclusion_reason_column: str,
        selection_criteria: SelectionConfig | None = None,
    ) -> WorkbookMetadata:
        try:
            workbook = load_workbook(source)
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            OSError,
            ValueError,
        ) as error:
            raise WorkbookValidationError(
                "invalid_workbook", "Uploaded file is not a valid XLSX workbook"
            ) from error
        try:
            selected_worksheet = worksheet_name or workbook.sheetnames[0]
            if selected_worksheet not in workbook.sheetnames:
                raise WorkbookValidationError(
                    "missing_worksheet", f"Worksheet not found: {selected_worksheet}"
                )
            worksheet = workbook[selected_worksheet]
            header_row, headers = _headers(worksheet)
            exclusion_reason_column = _resolve_exclusion_reason_column(
                headers, exclusion_reason_column
            )
            cls._validate_import_columns(
                headers, field_mappings, decision_column, exclusion_reason_column
            )
            record_count = _record_count(worksheet, header_row, len(headers))
            for column in (decision_column, exclusion_reason_column):
                if column not in headers:
                    headers.append(column)
                    worksheet.cell(header_row, len(headers), column)
            metadata = WorkbookMetadata(
                name=destination.name,
                worksheetName=selected_worksheet,
                headerRow=header_row,
                fieldMappings=field_mappings,
                decisionColumn=decision_column,
                exclusionReasonColumn=exclusion_reason_column,
                recordCount=record_count,
                selectionCriteria=selection_criteria,
            )
            cls._validate_decisions(worksheet, metadata, headers)
            destination.parent.mkdir(parents=True, exist_ok=True)
            cls._publish_workbook(workbook, destination)
            _write_json_atomic(
                metadata_path(destination), metadata.model_dump(by_alias=True)
            )
            return metadata
        finally:
            workbook.close()

    @staticmethod
    def _validate_import_columns(
        headers: list[str],
        field_mappings: dict[str, str],
        decision_column: str,
        exclusion_reason_column: str,
    ) -> None:
        from mapwisefox.web.api.workbooks._cache import (
            _mapped_columns,
            _validate_field_mappings,
        )

        _validate_field_mappings(headers, field_mappings)
        if decision_column == exclusion_reason_column:
            raise WorkbookValidationError(
                "duplicate_screening_columns",
                "Screening columns must have different names",
            )
        if {decision_column, exclusion_reason_column} & _mapped_columns(field_mappings):
            raise WorkbookValidationError(
                "screening_evidence_column_collision",
                "Screening columns must not replace mapped evidence fields",
            )

    @staticmethod
    def _validate_decisions(
        worksheet, metadata: WorkbookMetadata, headers: list[str]
    ) -> None:
        decision_index = headers.index(metadata.decision_column) + 1
        for offset in range(metadata.record_count):
            value = worksheet.cell(
                metadata.header_row + offset + 1, decision_index
            ).value
            normalized = "" if _is_blank(value) else str(value).strip().lower()
            if normalized not in DECISION_VALUES:
                raise WorkbookValidationError(
                    "invalid_decision",
                    f"Invalid decision value: {value}",
                    row=metadata.header_row + offset + 1,
                    column=metadata.decision_column,
                )

    @staticmethod
    def _publish_workbook(workbook, destination: Path) -> None:
        descriptor, temporary_name = tempfile.mkstemp(
            dir=destination.parent, suffix=".xlsx"
        )
        os.close(descriptor)
        try:
            workbook.save(temporary_name)
            validation = load_workbook(temporary_name, read_only=True)
            validation.close()
            os.replace(temporary_name, destination)
        finally:
            Path(temporary_name).unlink(missing_ok=True)

    def get(self, record_index: int) -> ScreeningRecord:
        self._validate_index(record_index)
        workbook = load_workbook(self.path, data_only=False)
        try:
            worksheet = workbook[self.metadata.worksheet_name]
            headers = self._current_headers(worksheet)
            values = {
                header: worksheet.cell(
                    self.metadata.header_row + record_index + 1, column
                ).value
                for column, header in enumerate(headers, start=1)
            }
            return self._to_record(record_index, values)
        finally:
            workbook.close()

    def undecided_indexes(self) -> list[int]:
        workbook = load_workbook(self.path, read_only=True, data_only=False)
        try:
            worksheet = workbook[self.metadata.worksheet_name]
            headers = self._current_headers(worksheet)
            decision_column_index = headers.index(self.metadata.decision_column) + 1
            blank_row_indices = [
                index
                for index, row in enumerate(
                    worksheet.iter_rows(
                        min_row=self.metadata.header_row + 1,
                        max_row=self.metadata.header_row + self.metadata.record_count,
                        min_col=decision_column_index,
                        max_col=decision_column_index,
                        values_only=True,
                    )
                )
                if _is_blank(row[0])
            ]
            return blank_row_indices
        finally:
            workbook.close()

    def update(
        self,
        record_index: int,
        decision: Decision,
        exclusion_reasons: list[str],
    ) -> ScreeningRecord:
        self._validate_index(record_index)
        workbook = load_workbook(self.path)
        try:
            worksheet = workbook[self.metadata.worksheet_name]
            row = self.metadata.header_row + record_index + 1
            decision_col_index = self._decision_col_index(worksheet)
            exclusion_reason_index = self._exclusion_reason_col_index(worksheet)
            worksheet.cell(row, decision_col_index, PERSISTED_DECISIONS[decision])
            worksheet.cell(
                row, exclusion_reason_index, LIST_SEPARATOR.join(exclusion_reasons)
            )
            self._publish_workbook(workbook, self.path)
        finally:
            workbook.close()
        return self.get(record_index)

    def delete(self) -> None:
        if not self.path.is_file():
            raise FileNotFoundError(self.path.name)
        self.path.unlink()
        metadata_path(self.path).unlink(missing_ok=True)

    def _current_headers(self, worksheet) -> list[str]:
        return [
            str(worksheet.cell(self.metadata.header_row, column).value).strip()
            for column in range(1, worksheet.max_column + 1)
        ]

    def _decision_col_index(self, worksheet) -> int:
        headers = self._current_headers(worksheet)
        return headers.index(self.metadata.decision_column) + 1

    def _exclusion_reason_col_index(self, worksheet) -> int:
        headers = self._current_headers(worksheet)
        column = _resolve_exclusion_reason_column(
            headers, self.metadata.exclusion_reason_column
        )
        return headers.index(column) + 1

    def _to_record(self, record_index: int, values: dict[str, Any]) -> ScreeningRecord:
        from mapwisefox.web.api.workbooks._cache import (
            SUPPORTED_FIELDS,
            _evidence_field,
            _mapped_value,
        )

        decision_value = values[self.metadata.decision_column]
        decision_key = (
            "" if _is_blank(decision_value) else str(decision_value).strip().lower()
        )
        reasons = Evidence._parse_list(
            {
                "reasons": _exclusion_reason_value(
                    values, self.metadata.exclusion_reason_column
                )
            },
            "reasons",
        )
        evidence_values = {
            _evidence_field(field): _mapped_value(
                values, field, self.metadata.field_mappings
            )
            for field in SUPPORTED_FIELDS
        }
        for field in ("doi", "title", "url"):
            evidence_values[field] = evidence_values[field] or ""
        evidence_values.update(
            cluster_id=record_index,
            include=decision_key == "include",
            exclude_reasons=reasons,
        )
        return ScreeningRecord(
            Evidence(**evidence_values), DECISION_VALUES[decision_key], reasons
        )

    def _validate_index(self, record_index: int) -> None:
        if record_index < 0 or record_index >= self.metadata.record_count:
            raise IndexError(record_index)
